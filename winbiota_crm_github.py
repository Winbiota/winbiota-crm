import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, os, json, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

creds_json = json.loads(os.environ['GOOGLE_CREDENTIALS'])
SHEET_ID   = os.environ['SHEET_ID']
EMAIL_USER = os.environ['EMAIL_USER']
EMAIL_PASS = os.environ['EMAIL_PASS']
OUTPUT     = '/tmp/Winbiota_CRM_Report.xlsx'

scopes = ['https://www.googleapis.com/auth/spreadsheets.readonly']
creds  = Credentials.from_service_account_info(creds_json, scopes=scopes)
sheets = build('sheets', 'v4', credentials=creds)
result = sheets.spreadsheets().values().get(
    spreadsheetId=SHEET_ID,
    range='CRM Winbiota',
    valueRenderOption='FORMATTED_VALUE',
    dateTimeRenderOption='FORMATTED_STRING'
).execute()
all_values = result.get('values', [])
max_cols = max(len(r) for r in all_values)
all_values = [r + [''] * (max_cols - len(r)) for r in all_values]
df_raw = pd.DataFrame(all_values)

data = df_raw.iloc[3:].copy()
data.columns = range(len(data.columns))
data = data.replace('', pd.NA)
data = data[~data[2].astype(str).str.contains('dummy data', na=False)]
data = data[data[2].astype(str).str.strip().replace('nan','') != ''].reset_index(drop=True)
N = len(data)

def parse_date(series):
    def _parse(val):
        v = str(val).strip()
        if v in ('', 'nan', 'None', 'NaT', '<NA>'): return pd.NaT
        for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%m/%d/%Y'):
            try: return datetime.datetime.strptime(v, fmt)
            except: pass
        return pd.to_datetime(v, errors='coerce')
    return series.apply(_parse)

estatus1 = data[12].astype(str).str.strip().str.upper()
estatus2 = data[16].astype(str).str.strip().str.upper()
contesto1 = data[11].astype(str).str.strip()
contesto2 = data[15].astype(str).str.strip()
comunic   = data[9].astype(str).str.strip()

buenos_vals1 = ['CONTESTO','CONTESTA','INTERES','LLAMADA','NO INT','SI']
buenos_vals2 = ['SI CONTESTA','INTERESADA']
buenos1 = estatus1.isin(buenos_vals1)
buenos2 = estatus2.isin(buenos_vals2)

# KPIs calculados en Python
total_leads       = N
contest_whats     = contesto1.replace('nan','').replace('','').replace('<NA>','').notna().sum()
# Contestaron whats = col 8 (Contestaron al Whats)
col8 = data[8].astype(str).str.strip()
contest_whats     = col8[~col8.isin(['','nan','None','<NA>'])].count()
ll1_hechas        = contesto1[~contesto1.isin(['','nan','None','<NA>'])].count()
ll1_buenos        = buenos1.sum()
ll2_hechas        = contesto2[~contesto2.isin(['','nan','None','<NA>'])].count()
ll2_buenos        = buenos2.sum()

# Bloqueo economico
keywords = ['PRECIO','CARO','VISTO','DINERO','PAGA']
excl = 'SIN PRECIO'
precio_mask = comunic.apply(lambda x: any(k in x.upper() for k in keywords) and excl not in x.upper())
bloqueo = precio_mask.sum()

pct = lambda a,b: round(a/b*100,1) if b>0 else 0

# Fechas de llamada desde cols auxiliares W(22) y X(23)
fecha_ll1 = parse_date(data[22]).dt.date
fecha_ll2 = parse_date(data[23]).dt.date
fecha_lead = parse_date(data[1]).dt.date

ll1_day        = fecha_ll1.dropna().value_counts().sort_index()
ll2_day        = fecha_ll2.dropna().value_counts().sort_index()
ll1_buenos_day = fecha_ll1[buenos1].dropna().value_counts().sort_index()
ll2_buenos_day = fecha_ll2[buenos2].dropna().value_counts().sort_index()
leads_day      = fecha_lead.dropna().value_counts().sort_index()
data['semana'] = parse_date(data[1]).dt.to_period('W')
leads_semana   = data.groupby('semana').size()

print(f"N={N}, ll1={ll1_hechas}, buenos1={ll1_buenos}, ll2={ll2_hechas}, buenos2={ll2_buenos}")
print(f"Fechas LL1={fecha_ll1.notna().sum()}, LL2={fecha_ll2.notna().sum()}")

# Hoja Datos
sel = data[[1,2,4,8,9,10,11,12,14,15,16,20]].copy()
sel.columns = ['Fecha','Nombre','Tel','Contestaron Whats','Comunicacion',
               'Fecha 1era Llamada','Contesto Llamada 1','Estatus 1era Llamada',
               'Fecha 2nda Llamada','Contesto Llamada 2','Estatus 2nda Llamada','Nota']
sel['Tel'] = sel['Tel'].astype(str).str.replace('p:','',regex=False).str.strip().replace('nan','')
sel['Fecha'] = parse_date(sel['Fecha']).dt.strftime('%d/%m/%Y').replace('NaT','')
sel['Fecha 1era Llamada'] = parse_date(data[22]).dt.strftime('%d/%m/%Y').replace('NaT','')
sel['Fecha 2nda Llamada'] = parse_date(data[23]).dt.strftime('%d/%m/%Y').replace('NaT','')
sel = sel.fillna('')

G_D='1B5E20'; G_M='2E7D32'; G_L='E8F5E9'; G_LL='F1F8E9'
WHITE='FFFFFF'; GRAY='F5F5F5'; AMBER='FFF8E1'; BLUE_D='1565C0'

def st(cell, bold=False, sz=9, col="000000", bg=None, ha='left', va='center', wrap=False):
    cell.font = Font(name='Arial', bold=bold, size=sz, color=col)
    if bg: cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)

thin  = Side(style='thin', color='C8E6C9')
bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
thin2 = Side(style='thin', color='B0BEC5')
bdr2  = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

def hrow(ws, row, labels, widths, bg=G_M):
    for i,(h,w) in enumerate(zip(labels,widths),1):
        c = ws.cell(row=row, column=i, value=h)
        st(c, bold=True, sz=9, col=WHITE, bg=bg, ha='center', wrap=True)
        c.border = bdr
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

def trow(ws, row, text, ncols, bg=G_D):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    st(c, bold=True, sz=12, col=WHITE, bg=bg, ha='center')
    ws.row_dimensions[row].height = 26

def section(ws, row, text, ncols, bg=BLUE_D):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    st(c, bold=True, sz=10, col=WHITE, bg=bg, ha='left')
    ws.row_dimensions[row].height = 20

def kpi(ws, row, value, label, fmt='number', bg=WHITE, ncols=10):
    c = ws.cell(row=row, column=1, value=value)
    if fmt == 'pct':
        c.number_format = '0.0"%"'
    st(c, bold=True, sz=13, col=G_D, bg=bg, ha='center')
    c.border = bdr2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    st(ws.cell(row=row, column=2, value=label), sz=10, bg=bg)
    ws.row_dimensions[row].height = 22

wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "Datos"
NCOLS1 = 12
trow(ws1, 1, 'WINBIOTA - Seguimiento Leads CRM', NCOLS1)
h1 = ['Fecha','Nombre','Telefono','Contestaron Whats','Comunicacion',
      'Fecha 1era Llamada','Contesto Llamada 1','Estatus 1era Llamada',
      'Fecha 2nda Llamada','Contesto Llamada 2','Estatus 2nda Llamada','Nota']
w1 = [12,26,14,12,22,13,10,20,13,10,20,26]
hrow(ws1, 2, h1, w1)
for ri,(_, row) in enumerate(sel.iterrows(), 3):
    bg = G_L if ri%2==0 else WHITE
    for ci, val in enumerate(row.values, 1):
        v = str(val) if str(val) not in ['','nan','NaT','<NA>'] else ''
        c = ws1.cell(row=ri, column=ci, value=v)
        st(c, sz=9, bg=bg, wrap=(ci in [4,5,8,12]))
        c.border = bdr
    ws1.row_dimensions[ri].height = 16
ws1.freeze_panes = 'A3'

ws2 = wb.create_sheet("Estadisticas")
NCOLS2 = 10

trow(ws2, 1, 'WINBIOTA - Estadisticas & KPIs', NCOLS2)
for i,w in enumerate([14,30,12,12,12,11,11,13,14,13],1):
    ws2.column_dimensions[get_column_letter(i)].width = w

r = 3
section(ws2, r, 'TOTALES', NCOLS2); r+=1

# Escribir KPIs directamente como numeros
rows_totales = [
    (total_leads,    'Total leads',                         G_LL),
    (contest_whats,  'Contestaron WhatsApp',                WHITE),
    (ll1_hechas,     'Llamadas 1 realizadas',               G_LL),
    (ll1_buenos,     'Llamada 1 efectiva (buenos estatus)', WHITE),
    (ll2_hechas,     'Llamadas 2 realizadas',               G_LL),
    (ll2_buenos,     'Llamada 2 efectiva (buenos estatus)', WHITE),
    (bloqueo,        'Piden precio / bloqueo economico',    AMBER),
]
for val, label, bg in rows_totales:
    kpi(ws2, r, int(val), label, 'number', bg, NCOLS2); r+=1
r+=1

section(ws2, r, 'PORCENTAJES', NCOLS2); r+=1
pct_rows = [
    (pct(contest_whats, total_leads),                    'Contestaron Whats / Total leads',                      GRAY),
    (pct(ll1_hechas, contest_whats),                     'Llamadas 1 hechas / Contestaron Whats',                WHITE),
    (pct(ll1_buenos, ll1_hechas),                        'Llamada 1 efectiva (buenos) / Llamadas 1 hechas',      GRAY),
    (pct(ll2_hechas, ll1_hechas),                        'Llamadas 2 hechas / Llamadas 1 hechas',                WHITE),
    (pct(ll2_buenos, ll2_hechas),                        'Llamada 2 efectiva (buenos) / Llamadas 2 hechas',      GRAY),
    (pct(ll1_buenos+ll2_buenos, ll1_hechas+ll2_hechas),  '% Efectividad total',                                  WHITE),
    (pct(bloqueo, total_leads),                          '% leads bloqueo economico',                            AMBER),
]
for val, label, bg in pct_rows:
    c = ws2.cell(row=r, column=1, value=f"{val}%")
    st(c, bold=True, sz=13, col=G_D, bg=bg, ha='center')
    c.border = bdr2
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS2)
    st(ws2.cell(row=r, column=2, value=label), sz=10, bg=bg)
    ws2.row_dimensions[r].height = 22; r+=1
r+=1

section(ws2, r, 'LLAMADAS POR DIA', NCOLS2); r+=1
all_d = sorted(set(list(ll1_day.index)+list(ll2_day.index)))
if not all_d:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS2)
    c = ws2.cell(row=r, column=1, value='Sin fechas de llamada registradas')
    st(c, sz=9, col='888888', bg=GRAY, ha='center')
    ws2.row_dimensions[r].height = 18; r+=2
else:
    hrow(ws2, r,
         ['Fecha','LL1 hechas','LL1 buenos','% Efect 1','LL2 hechas','LL2 buenos','% Efect 2',
          'LL tot hechas','Efectivas tot','% Efect tot'],
         [13,10,10,10,10,10,10,12,13,12]); r+=1
    tot_ll1=tot_ll2=tot_b1=tot_b2=0
    for date in all_d:
        bg  = G_L if r%2==0 else WHITE
        ll1 = ll1_day.get(date,0); ll2 = ll2_day.get(date,0)
        b1  = ll1_buenos_day.get(date,0); b2 = ll2_buenos_day.get(date,0)
        tot_ll1+=ll1; tot_ll2+=ll2; tot_b1+=b1; tot_b2+=b2
        try: ds = datetime.datetime.strptime(str(date),'%Y-%m-%d').strftime('%d/%m/%Y')
        except: ds = str(date)
        vals = [ds, ll1, b1, f"{pct(b1,ll1)}%", ll2, b2, f"{pct(b2,ll2)}%"]
        for ci,val in enumerate(vals,1):
            c = ws2.cell(row=r, column=ci, value=val)
            st(c, sz=9, bg=bg, ha='center' if ci>1 else 'left'); c.border = bdr2
        tot_d=ll1+ll2; btot_d=b1+b2
        for ci,val in enumerate([tot_d, btot_d, f"{pct(btot_d,tot_d)}%"], 8):
            c = ws2.cell(row=r, column=ci, value=val)
            st(c, bold=True, sz=9, bg=bg, ha='center'); c.border = bdr2
        ws2.row_dimensions[r].height = 16; r+=1
    tot_d=tot_ll1+tot_ll2; btot=tot_b1+tot_b2
    vals_tot = ['TOTAL', tot_ll1, tot_b1, f"{pct(tot_b1,tot_ll1)}%",
                tot_ll2, tot_b2, f"{pct(tot_b2,tot_ll2)}%",
                tot_d, btot, f"{pct(btot,tot_d)}%"]
    for ci,val in enumerate(vals_tot,1):
        c = ws2.cell(row=r, column=ci, value=val)
        st(c, bold=True, sz=10, col=WHITE, bg=G_M, ha='center'); c.border = bdr2
    ws2.row_dimensions[r].height = 20; r+=2

section(ws2, r, 'LEADS NUEVOS POR DIA', NCOLS2); r+=1
hrow(ws2, r, ['Fecha','Leads nuevos']+['']*8, [13,12]+[12]*8); r+=1
total_dia=0
for date in sorted(leads_day.index):
    bg = G_L if r%2==0 else WHITE
    count = leads_day[date]; total_dia += count
    try: ds = datetime.datetime.strptime(str(date),'%Y-%m-%d').strftime('%d/%m/%Y')
    except: ds = str(date)
    c1 = ws2.cell(row=r, column=1, value=ds)
    st(c1, sz=9, bg=bg); c1.border = bdr2
    c2 = ws2.cell(row=r, column=2, value=int(count))
    st(c2, sz=10, bg=bg, ha='center'); c2.border = bdr2
    for ci in range(3, NCOLS2+1):
        ws2.cell(row=r, column=ci).fill = PatternFill('solid', start_color=bg)
    ws2.row_dimensions[r].height = 16; r+=1
for ci,val in enumerate(['TOTAL',int(total_dia)]+['']*8,1):
    c = ws2.cell(row=r, column=ci, value=val)
    st(c, bold=True, sz=10, col=WHITE, bg=G_M, ha='center'); c.border = bdr2
ws2.row_dimensions[r].height = 20; r+=2

section(ws2, r, 'LEADS NUEVOS POR SEMANA', NCOLS2); r+=1
hrow(ws2, r, ['Semana','Leads nuevos']+['']*8, [20,12]+[12]*8); r+=1
total_sem=0
for period, count in leads_semana.items():
    bg = G_L if r%2==0 else WHITE
    try: label = f"{period.start_time.strftime('%d/%m/%Y')} - {period.end_time.strftime('%d/%m/%Y')}"
    except: label = str(period)
    c1 = ws2.cell(row=r, column=1, value=label)
    st(c1, sz=9, bg=bg); c1.border = bdr2
    c2 = ws2.cell(row=r, column=2, value=int(count))
    st(c2, sz=10, bg=bg, ha='center'); c2.border = bdr2
    for ci in range(3, NCOLS2+1):
        ws2.cell(row=r, column=ci).fill = PatternFill('solid', start_color=bg)
    total_sem += count
    ws2.row_dimensions[r].height = 16; r+=1
for ci,val in enumerate(['TOTAL',int(total_sem)]+['']*8,1):
    c = ws2.cell(row=r, column=ci, value=val)
    st(c, bold=True, sz=10, col=WHITE, bg=G_M, ha='center'); c.border = bdr2
ws2.row_dimensions[r].height = 20

wb.save(OUTPUT)
print(f"Listo. N={N}")

now = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
msg = MIMEMultipart()
msg['From'] = EMAIL_USER
msg['To']   = EMAIL_USER
msg['Subject'] = f'Winbiota CRM - Reporte {now}'
msg.attach(MIMEText(f"Reporte CRM generado el {now}.\nLeads: {N} | LL1 hechas: {ll1_hechas} | Buenos1: {ll1_buenos} | LL2 hechas: {ll2_hechas} | Buenos2: {ll2_buenos}", 'plain'))
with open(OUTPUT, 'rb') as f:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(f.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', f'attachment; filename="Winbiota_CRM_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"')
msg.attach(part)
with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
    server.login(EMAIL_USER, EMAIL_PASS)
    server.sendmail(EMAIL_USER, EMAIL_USER, msg.as_string())
print("Email enviado.")
